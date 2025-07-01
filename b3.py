from openpyxl import load_workbook
w = load_workbook("sales.xlsx").active
data = list(w.iter_rows(values_only=True))[1:]
print("Total Sales :",sum(i[2]*i[1] for i in data))
print("Total Quantity :",sum(i[1] for i in data))
print("Avg Price per product :",sum(i[2]*i[1] for i in data)/sum(i[1] for i in data))
print("Most Expensive Product :",sorted(data,key=lambda x:x[2])[-1][0])
print("Best Selling Product :",sorted(data,key=lambda x:x[1])[-1][0])
